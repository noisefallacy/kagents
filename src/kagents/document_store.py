"""Document loading and search primitives for the MVP."""

from __future__ import annotations

from dataclasses import dataclass
from email import policy
from email.parser import BytesParser
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document


SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx", ".eml"}


@dataclass(frozen=True)
class SearchHit:
    path: str
    score: int
    excerpt: str


def iter_supported_files(root: Path) -> Iterable[Path]:
    if not root.exists():
        return []

    return (
        path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()

    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="ignore")

    if suffix == ".pdf":
        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    if suffix == ".docx":
        document = Document(str(path))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)

    if suffix == ".xlsx":
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if values:
                    lines.append(" | ".join(values))
        return "\n".join(lines)

    if suffix == ".eml":
        with path.open("rb") as handle:
            message = BytesParser(policy=policy.default).parse(handle)

        lines: list[str] = []
        for header_name in ("From", "To", "Subject", "Date"):
            header_value = message.get(header_name)
            if header_value:
                lines.append(f"{header_name}: {header_value}")

        body_parts: list[str] = []
        if message.is_multipart():
            for part in message.walk():
                if part.get_content_type() == "text/plain" and not part.get_filename():
                    body_parts.append(part.get_content())
        else:
            body_parts.append(message.get_content())

        lines.extend(part.strip() for part in body_parts if part and part.strip())
        return "\n".join(lines)

    raise ValueError(f"Unsupported file type: {path.suffix}")


def score_text(query: str, text: str) -> int:
    terms = [term for term in re.findall(r"\w+", query.lower()) if term]
    lowered_text = text.lower()
    return sum(lowered_text.count(term) for term in terms)


def make_excerpt(text: str, query: str, window: int = 220) -> str:
    compact_text = re.sub(r"\s+", " ", text).strip()
    if not compact_text:
        return ""

    query_terms = [term for term in re.findall(r"\w+", query.lower()) if term]
    first_index = min(
        (compact_text.lower().find(term) for term in query_terms if compact_text.lower().find(term) != -1),
        default=0,
    )
    start = max(first_index - window // 3, 0)
    end = min(start + window, len(compact_text))
    excerpt = compact_text[start:end].strip()

    if start > 0:
        excerpt = "..." + excerpt
    if end < len(compact_text):
        excerpt = excerpt + "..."
    return excerpt


def search_directory(root: Path, query: str, max_results: int = 5) -> list[SearchHit]:
    hits: list[SearchHit] = []

    for path in iter_supported_files(root):
        text = extract_text(path)
        score = score_text(query, text)
        if score <= 0:
            continue

        hits.append(
            SearchHit(
                path=str(path),
                score=score,
                excerpt=make_excerpt(text, query),
            )
        )

    hits.sort(key=lambda item: item.score, reverse=True)
    return hits[:max_results]
